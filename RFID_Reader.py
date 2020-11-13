# Read Excel
import openpyxl
import pytz
import datetime
current_time = datetime.datetime.now(pytz.timezone('Asia/Hong_Kong'))
print(current_time)
# Open Excel File
wb = openpyxl.load_workbook(r'C:\Users\Benedictho1215\Desktop\RFID test.xlsx')
# Check all the sheetnames in file
sheetnames = wb.sheetnames
print(sheetnames)
# Search according the sheet name
record = wb['Retrieve Record']
tool = wb['Tool']
print(tool)

def new_tool_RFID():
    new_id = input('New tool ID')
    new_name = input('New Tool name')
    New_Calibration = input('Calibration Date')
    try:
        int(new_id)
        id = int(new id);
        col1 = [A[0].value for A in tool]
        k = len(col1) + 1
        if New_ID in col1:
            print('The id is exist')
            print('Return to homepage')
    except ValueError: 
        print('Wrong ID')
        break;




def toolretrieve():
    id = input('tool id')
    name = input('tool name')
    try:
        int(id)
        id = int(id)
    except ValueError:
        print('Wrong tool id')
        break

def main():
    add_new_tool
    tool_retrieve
    print('Please choose your action, Adding new tool or Tool Retrieve')
    if input = Adding New Tool:
    new_tool_RFID()
    if input = Tool Retrieve:
    toolretrieve()
system = False
while system == False
    main()
    stop = input('Press enter to continue. Enter [end] to stop the programme')
    if stop == 'end':
        system = True
    else:
        system = False